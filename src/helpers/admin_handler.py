import json
import random

from datetime import date, datetime, time
from html import escape
from io import BytesIO

from aiogram.filters.command import CommandObject
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import DATETIME_FORMAT, UFA_TZ
from src.bot.texts import admin_texts
from src.condition import ConditionType
from src.database.models import Condition


def _condition_name(condition) -> str:
    try:
        return ConditionType[condition.action].value._name
    except Exception:
        return condition.action


def _giveaway_start_as_datetime_iso(giveaway_start_date: date) -> str:
    return datetime.combine(giveaway_start_date, time.min, tzinfo=UFA_TZ).isoformat()


def _classify_participants(giveaway) -> tuple[list, list, set[int], list]:
    participants = sorted(giveaway.participants or [], key=lambda participant: participant.created_at)
    conditions = giveaway.conditions or []
    mandatory_condition_ids = {condition.id for condition in conditions if condition.mandatory}
    if not mandatory_condition_ids:
        return participants, [], mandatory_condition_ids, conditions

    winnable_participants: list = []
    not_winnable_participants: list = []
    for participant in participants:
        records_map = {record.condition_id: record for record in (participant.records or [])}
        has_all_mandatory = all(records_map.get(condition_id) and records_map[condition_id].passed for condition_id in mandatory_condition_ids)
        if has_all_mandatory:
            winnable_participants.append(participant)
        else:
            not_winnable_participants.append(participant)
    return winnable_participants, not_winnable_participants, mandatory_condition_ids, conditions


def _format_participants_text(giveaway, *, show_all: bool) -> str:
    winnable_participants, not_winnable_participants, _, _ = _classify_participants(giveaway)
    shown_count = len(winnable_participants) + len(not_winnable_participants) if show_all else len(winnable_participants)
    lines = [
        f"<b>👥 Участники розыгрыша #{giveaway.id} — {escape(giveaway.name)}</b>",
        f"Всего участников: <b>{len(winnable_participants) + len(not_winnable_participants)}</b>",
        f"Проходящие (mandatory): <b>{len(winnable_participants)}</b> · Не проходящие: <b>{len(not_winnable_participants)}</b>",
        f"Режим: <b>{'все участники' if show_all else 'только проходящие'}</b>",
        f"Показано в режиме: <b>{shown_count}</b>",
    ]
    if shown_count == 0:
        lines.append("Подходящих участников в выбранном режиме пока нет.")
        return "\n\n".join(lines)
    lines.append("Детальный список выгружайте кнопкой <b>⬇️ Скачать Excel</b>.")
    return "\n\n".join(lines)


def _condition_extra_info(condition, record) -> str:
    if record is None:
        return "Нет данных"
    config = record.config or {}
    if condition.action == "website_order":
        codes = [str(value).strip() for value in config.get("order_codes", []) if str(value).strip()]
        return f"Коды заказов: {', '.join(codes)}" if codes else f"Выполнений: {record.complete}"
    if condition.action == "website_review":
        email = str(config.get("email") or "").strip()
        if not email:
            emails = [str(value).strip() for value in config.get("review_emails", []) if str(value).strip()]
            email = emails[-1] if emails else ""
        return f"Email: {email}" if email else f"Выполнений: {record.complete}"
    if condition.action == "ref_join":
        referrals = [str(value) for value in config.get("referrals", []) if str(value).strip()]
        if not referrals:
            return f"Приглашено: {record.complete}"
        return f"Приглашено: {len(referrals)} ({', '.join(referrals)})"
    if condition.action == "self_join":
        return f"Выполнений: {record.complete}"
    if config:
        return f"Выполнений: {record.complete}; Конфиг: {json.dumps(config, ensure_ascii=False)}"
    return f"Выполнений: {record.complete}"


def _participant_excel_row(participant, conditions: list, mandatory_condition_ids: set[int]) -> list[str]:
    user = participant.user
    user_id = user.id if user is not None else participant.user_id
    full_name = _participant_display_name(participant)
    phone = getattr(user, "phone", None) if user is not None else None

    joined_at = participant.created_at
    if joined_at.tzinfo is None:
        joined_at = joined_at.replace(tzinfo=UFA_TZ)
    joined_text = joined_at.astimezone(UFA_TZ).strftime(DATETIME_FORMAT)

    records_map = {record.condition_id: record for record in (participant.records or [])}
    passed_total = sum(1 for record in records_map.values() if record.passed)
    mandatory_passed = sum(1 for condition_id in mandatory_condition_ids if records_map.get(condition_id) and records_map[condition_id].passed)
    is_winnable = (len(mandatory_condition_ids) == 0) or (mandatory_passed == len(mandatory_condition_ids))
    accumulated_tickets = _participant_accumulated_tickets(participant, conditions)

    row: list[str] = [
        str(participant.id),
        str(user_id),
        full_name,
        phone or "",
        joined_text,
        "ДА" if is_winnable else "НЕТ",
        str(passed_total),
        str(len(conditions)),
        str(mandatory_passed),
        str(len(mandatory_condition_ids)),
        str(accumulated_tickets),
    ]

    for condition in sorted(conditions, key=lambda cond: cond.id):
        record = records_map.get(condition.id)
        row.append("ДА" if (record and record.passed) else "НЕТ")
        row.append(_condition_extra_info(condition, record))
    return row


def _participant_accumulated_tickets(participant, conditions: list[Condition]) -> int:
    records_map = {record.condition_id: record for record in (participant.records or [])}
    reward_tickets = 0
    passed_conditions = 0
    for condition in conditions:
        record = records_map.get(condition.id)
        if record is not None and record.passed:
            passed_conditions += 1
        if condition.reward is not None and record is not None:
            complete = max(int(record.complete or 0), 0)
            max_repeats = Condition.resolve_max_repeats(condition.action, condition.repeatable, condition.config, condition.required)
            counted = complete if max_repeats is None else min(complete, max_repeats)
            if counted > 0:
                reward_tickets += counted * int(condition.reward)

    base_ticket = 1 if len(conditions) == 0 or passed_conditions >= len(conditions) else 0
    return max(base_ticket + reward_tickets, 0)


def _prepare_sheet(ws, headers: list[str]) -> None:
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill


def _autosize_sheet(ws) -> None:
    max_width = 60
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), max_width)


def _build_participants_excel(giveaway) -> bytes:
    winnable_participants, not_winnable_participants, mandatory_condition_ids, conditions = _classify_participants(giveaway)
    wb = Workbook()
    ws_winnable = wb.active
    ws_winnable.title = "Проходящие"
    ws_not_winnable = wb.create_sheet("Не проходят")

    headers = [
        "ID участия",
        "Telegram ID",
        "Имя участника",
        "Телефон",
        "Дата участия (Уфа)",
        "Проходит обязательные условия",
        "Пройдено условий",
        "Всего условий",
        "Пройдено обязательных условий",
        "Всего обязательных условий",
        "Накопленные билеты",
    ]
    for condition in sorted(conditions, key=lambda cond: cond.id):
        condition_title = _condition_name(condition)
        headers.append(f"{condition_title} — пройдено")
        headers.append(f"{condition_title} — детали")

    _prepare_sheet(ws_winnable, headers)
    _prepare_sheet(ws_not_winnable, headers)

    for participant in winnable_participants:
        ws_winnable.append(_participant_excel_row(participant, conditions, mandatory_condition_ids))
    for participant in not_winnable_participants:
        ws_not_winnable.append(_participant_excel_row(participant, conditions, mandatory_condition_ids))

    _autosize_sheet(ws_winnable)
    _autosize_sheet(ws_not_winnable)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _participant_display_name(participant) -> str:
    user = participant.user
    user_id = user.id if user is not None else participant.user_id
    full_name = " ".join(
        [
            part
            for part in [getattr(user, "first_name", None), getattr(user, "last_name", None)]
            if part
        ]
    ).strip() if user is not None else ""
    if full_name:
        return full_name
    return f"user_{user_id}"


def _participant_ticket_weight(participant, conditions_by_id: dict[int, Condition]) -> int:
    tickets = 1
    for record in participant.records or []:
        condition = conditions_by_id.get(record.condition_id)
        if condition is None or condition.reward is None:
            continue
        complete = max(int(record.complete or 0), 0)
        max_repeats = Condition.resolve_max_repeats(condition.action, condition.repeatable, condition.config, condition.required)
        counted = complete if max_repeats is None else min(complete, max_repeats)
        if counted == 0:
            continue
        tickets += counted * int(condition.reward)
    return max(tickets, 1)


def _draw_winners(giveaway) -> tuple[list[dict], int]:
    from src.database.schemas.giveaway import _normalize_prizes

    winnable_participants, _, _, conditions = _classify_participants(giveaway)
    if not winnable_participants:
        return [], 0

    normalized_prizes = _normalize_prizes((giveaway.prizes or {}).items())
    prize_slots = sorted(normalized_prizes.items(), key=lambda item: int(item[0]))
    total_slots = sum(max(int(prize.amount), 0) for _, prize in prize_slots)
    if total_slots <= 0:
        return [], 0

    conditions_by_id = {condition.id: condition for condition in conditions}
    pool: list[tuple[object, int]] = [
        (participant, _participant_ticket_weight(participant, conditions_by_id))
        for participant in winnable_participants
    ]

    selected_at = datetime.now(tz=UFA_TZ).isoformat()
    winners: list[dict] = []
    for place, prize in prize_slots:
        amount = max(int(prize.amount), 0)
        for slot in range(1, amount + 1):
            if not pool:
                return winners, total_slots

            weights = [weight for _, weight in pool]
            selected_index = random.choices(range(len(pool)), weights=weights, k=1)[0]
            winner_participant, tickets = pool.pop(selected_index)
            winner_user = winner_participant.user
            winner_user_id = winner_user.id if winner_user is not None else winner_participant.user_id
            winners.append(
                {
                    "place": int(place),
                    "slot": slot,
                    "participant_id": int(winner_participant.id),
                    "user_id": int(winner_user_id),
                    "full_name": _participant_display_name(winner_participant),
                    "phone": getattr(winner_user, "phone", None) if winner_user is not None else None,
                    "prize_name": prize.name,
                    "tickets": int(tickets),
                    "selected_at": selected_at,
                }
            )
    return winners, total_slots


def _format_winners_text(giveaway) -> str:
    winners = [winner for winner in (giveaway.winners or []) if isinstance(winner, dict)]
    lines = [f"<b>🏆 Победители розыгрыша #{giveaway.id} — {escape(giveaway.name)}</b>"]
    if not winners:
        lines.append(admin_texts.Winners.not_found)
        return "\n\n".join(lines)

    lines.append(f"Всего определено победителей: <b>{len(winners)}</b>")
    places: dict[int, list[dict]] = {}
    for winner in winners:
        try:
            place = int(winner.get("place"))
        except Exception:
            continue
        places.setdefault(place, []).append(winner)

    for place in sorted(places):
        lines.append(f"<b>{place} место</b>")
        for index, winner in enumerate(places[place], start=1):
            user_id = winner.get("user_id")
            user_display = "?" if user_id is None else str(user_id)
            full_name = escape(str(winner.get("full_name") or f"user_{user_display}"))
            tickets = winner.get("tickets")
            prize_name = escape(str(winner.get("prize_name") or "Приз"))
            ticket_suffix = f", билетов: {tickets}" if isinstance(tickets, int) else ""
            lines.append(f"{index}. {full_name} · ID: <code>{user_display}</code>{ticket_suffix} · {prize_name}")
    return "\n".join(lines)


def _participants_export_caption(giveaway, passed_count: int, all_count: int) -> str:
    export_time = datetime.now(tz=UFA_TZ).strftime(DATETIME_FORMAT)
    safe_name = escape(giveaway.name)
    return (
        f"<b>{safe_name}\n📊Участники розыгрыша</b>: {passed_count}/{all_count}\n"
        f"Выгрузка от {export_time}"
    )


def _iter_prizes(giveaway) -> list[tuple[int, str, int]]:
    raw_prizes = giveaway.prizes or {}
    if not isinstance(raw_prizes, dict):
        try:
            raw_prizes = dict(raw_prizes)
        except Exception:
            return []

    parsed: list[tuple[int, str, int]] = []
    for raw_place, raw_prize in raw_prizes.items():
        try:
            place = int(raw_place)
        except (TypeError, ValueError):
            continue

        if isinstance(raw_prize, dict):
            name = str(raw_prize.get("name") or "").strip()
            amount_raw = raw_prize.get("amount", 0)
        else:
            name = str(getattr(raw_prize, "name", "") or "").strip()
            amount_raw = getattr(raw_prize, "amount", 0)

        try:
            amount = int(amount_raw)
        except (TypeError, ValueError):
            amount = 0

        if not name:
            name = f"Приз #{place}"
        parsed.append((place, name, max(amount, 0)))

    parsed.sort(key=lambda value: value[0])
    return parsed


def _prize_for_place(giveaway, place: int) -> tuple[str, int] | None:
    for prize_place, prize_name, prize_amount in _iter_prizes(giveaway):
        if prize_place == place:
            return prize_name, prize_amount
    return None


def _format_prizes_block(giveaway) -> str:
    prizes = _iter_prizes(giveaway)
    lines = ["ЧТО МОЖНО ВЫИГРАТЬ"]
    if not prizes:
        lines.append("Призы пока не настроены.")
        return "\n".join(lines)
    for place, name, amount in prizes:
        lines.append(f"{place}. {escape(name)} x{amount}шт.")
    return "\n".join(lines)


def _participants_passed_all_conditions(giveaway) -> list:
    participants = sorted(giveaway.participants or [], key=lambda participant: participant.created_at)
    conditions = giveaway.conditions or []
    if not conditions:
        return participants

    passed: list = []
    for participant in participants:
        records_map = {record.condition_id: record for record in (participant.records or [])}
        if all(records_map.get(condition.id) and records_map[condition.id].passed for condition in conditions):
            passed.append(participant)
    return passed


def _pick_weighted_random_participant(giveaway, *, excluded_user_ids: set[int]) -> tuple[int, int] | None:
    passed_participants = _participants_passed_all_conditions(giveaway)
    if not passed_participants:
        return None

    conditions_by_id = {int(condition.id): condition for condition in (giveaway.conditions or [])}
    pool: list[tuple[int, int]] = []
    for participant in passed_participants:
        user_id = int(participant.user_id)
        if user_id in excluded_user_ids:
            continue
        tickets = _participant_ticket_weight(participant, conditions_by_id)
        pool.append((user_id, max(int(tickets), 1)))

    if not pool:
        return None

    weights = [tickets for _, tickets in pool]
    selected_index = random.choices(range(len(pool)), weights=weights, k=1)[0]
    return pool[selected_index]


def _decide_winner_place_prompt(giveaway) -> str:
    return f"{admin_texts.DecideWinner.place_prompt}\n\n{_format_prizes_block(giveaway)}"


def _parse_command_giveaway_id(command: CommandObject | None) -> int | None:
    if command is None or not command.args:
        return None

    raw_value = command.args.strip().split()[0]
    try:
        giveaway_id = int(raw_value)
    except (TypeError, ValueError):
        return None

    if giveaway_id <= 0:
        return None
    return giveaway_id


def _normalize_winners(raw_winners) -> list[dict]:
    if not isinstance(raw_winners, list):
        return []
    return [dict(value) for value in raw_winners if isinstance(value, dict)]


def _winners_user_ids_for_other_places(winners: list[dict], *, place: int) -> set[int]:
    user_ids: set[int] = set()
    for winner in winners:
        try:
            winner_place = int(winner.get("place"))
            user_id = int(winner.get("user_id"))
        except Exception:
            continue
        if winner_place == place:
            continue
        user_ids.add(user_id)
    return user_ids


def _save_winner(
    giveaway,
    *,
    place: int,
    user_id: int,
    prize_name: str,
    selected_by_admin_id: int,
    source: str,
    tickets: int | None = None,
) -> None:
    winners = _normalize_winners(giveaway.winners)
    new_winner = {
        "place": place,
        "user_id": user_id,
        "prize_name": prize_name,
        "selected_at": datetime.now(tz=UFA_TZ).isoformat(),
        "selected_by_admin_id": selected_by_admin_id,
        "selected_source": source,
        "phone_shared": False,
    }
    if tickets is not None:
        try:
            normalized_tickets = int(tickets)
        except (TypeError, ValueError):
            normalized_tickets = None
        if normalized_tickets is not None and normalized_tickets > 0:
            new_winner["tickets"] = normalized_tickets

    updated_winners: list[dict] = []
    for winner in winners:
        try:
            winner_place = int(winner.get("place", -1))
        except Exception:
            updated_winners.append(winner)
            continue
        if winner_place != place:
            updated_winners.append(winner)
    updated_winners.append(new_winner)
    updated_winners.sort(
        key=lambda winner: int(winner.get("place", 10**9)) if str(winner.get("place", "")).isdigit() else 10**9
    )
    giveaway.winners = updated_winners


__all__ = [
    "_build_participants_excel",
    "_classify_participants",
    "_decide_winner_place_prompt",
    "_draw_winners",
    "_format_participants_text",
    "_format_prizes_block",
    "_format_winners_text",
    "_giveaway_start_as_datetime_iso",
    "_normalize_winners",
    "_parse_command_giveaway_id",
    "_participants_export_caption",
    "_pick_weighted_random_participant",
    "_prize_for_place",
    "_save_winner",
    "_winners_user_ids_for_other_places",
]
